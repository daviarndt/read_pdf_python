import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime
import mysql.connector

def execute_insert(cursor, invoice_number, invoice_date, file_name, status):
    sql = "INSERT INTO invoice_records (invoice_number, invoice_date, file_name,status) VALUES (%s, %s, %s, %s)"
    val = (invoice_number, invoice_date, file_name,status)
    cursor.execute(sql, val)

def main():
    # STARTUP

    # Database Connection
    db = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="process_invoices"
    )
    cursor = db.cursor()
    print("--- Successfully connected to database... ---")

    # Get files from directory
    directory = 'pdf_invoices'
    files = os.listdir(directory)
    files_quantity = len(files)

    if files_quantity == 0:
        raise Exception("No files found in the directory")

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice Imports'

    ws['A1'] = 'Invoice #'
    ws['B1'] = 'Date'
    ws['C1'] = 'File Name'
    ws['D1'] = 'Status'

    last_empty_line = 1
    while ws["D" + str(last_empty_line)].value is not None:
        last_empty_line += 1

    # WORK
    for file in files:
        try:
            with pdfplumber.open(directory + "/" + file) as pdf:
                first_page = pdf.pages[0]
                pdf_text = first_page.extract_text()

            inv_number_re_pattern = r'INVOICE #(\d+)'
            inv_date_re_pattern = r'DATE (\d{2}/\d{2}/\d{4})'

            match_number = re.search(inv_number_re_pattern, pdf_text)
            match_date = re.search(inv_date_re_pattern, pdf_text)

            if match_number:
                ws['A{}'.format(last_empty_line)] = match_number.group(1)
            else:
                raise Exception("Couldn't find invoice number")

            if match_date:
                ws['B{}'.format(last_empty_line)] = match_date.group(1)
            else:
                raise Exception("Couldn't find invoice date")

            ws['C{}'.format(last_empty_line)] = file
            ws['D{}'.format(last_empty_line)] = "Completed"

            execute_insert(cursor, match_number.group(1), match_date.group(1), file, "Completed")
            db.commit()

            last_empty_line += 1

        except Exception as e:
            print(f"Error processing file: {e}")

            ws['C{}'.format(last_empty_line)] = file
            ws['D{}'.format(last_empty_line)] = "Exception: {}".format(e)

            execute_insert(cursor, "N/A", "N/A", file, "Exception: {}".format(e))
            db.commit()

            last_empty_line += 1

    cursor.close()
    db.close()

    full_now = str(datetime.now()).replace(":", "-")
    dot_index = full_now.index(".")
    now = full_now[:dot_index]
    wb.save("Invoices - {}.xlsx".format(now))

if __name__ == "__main__":
    main()